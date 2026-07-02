import io
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

LOGIN_URL = '/system/'
PER_PAGE  = 20


def _paginate(qs, page, per_page=PER_PAGE):
    from django.core.paginator import Paginator
    p = Paginator(qs, per_page)
    return p.get_page(page), p.num_pages


def _excel_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return resp


def _style_header(ws, headers):
    fill  = PatternFill('solid', fgColor='C8920A')
    font  = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    align = Alignment(horizontal='center', vertical='center')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill; c.font = font; c.alignment = align
    ws.row_dimensions[1].height = 26


def _date_filters(request):
    return request.GET.get('date_from', ''), request.GET.get('date_to', '')


# ─── MAIN HUB ───────────────────────────────────────────────────────────────
@login_required(login_url=LOGIN_URL)
def reports_home(request):
    from invoices.models import Invoice
    from clients.models import Client
    from employees.models import Employee
    from inventory.models import Item

    qs = Invoice.objects.exclude(status='draft')
    stats = {
        'total_revenue':  qs.aggregate(s=Sum('total_amount'))['s'] or 0,
        'total_paid':     qs.filter(status='fully_paid').aggregate(s=Sum('total_amount'))['s'] or 0,
        'total_vat':      qs.aggregate(s=Sum('vat_amount'))['s'] or 0,
        'invoices_count': qs.count(),
        'clients_count':  Client.objects.count(),
        'employees_count':Employee.objects.count(),
        'items_count':    Item.objects.count(),
    }
    return render(request, 'reports/reports_home.html', {'stats': stats})


# ─── 1. INVOICES ────────────────────────────────────────────────────────────
@login_required(login_url=LOGIN_URL)
def report_invoices(request):
    from invoices.models import Invoice
    date_from, date_to = _date_filters(request)
    status = request.GET.get('status', '')
    export = request.GET.get('export', '')
    page   = int(request.GET.get('page', 1))
    STATUS_MAP = {'draft':'مسودة','final':'نهائية','fully_paid':'مدفوعة','partial_paid':'مدفوعة جزئياً','cancelled':'ملغاة'}

    qs = Invoice.objects.select_related('client','project').order_by('-issue_date')
    if date_from: qs = qs.filter(issue_date__gte=date_from)
    if date_to:   qs = qs.filter(issue_date__lte=date_to)
    if status:    qs = qs.filter(status=status)

    if export == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'الفواتير'
        _style_header(ws, ['#','رقم الفاتورة','العميل','المشروع','التاريخ','الحالة','قبل الضريبة','الضريبة','الإجمالي','المدفوع','المتبقي'])
        for i, inv in enumerate(qs, 1):
            ws.append([i, inv.invoice_number, inv.client.name if inv.client else '', inv.project.name if inv.project else '',
                str(inv.issue_date), STATUS_MAP.get(inv.status,''), float(inv.subtotal), float(inv.vat_amount),
                float(inv.total_amount), float(inv.paid_amount), float(inv.adjusted_remaining)])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 16
        return _excel_response(wb, 'تقرير_الفواتير')

    totals = qs.aggregate(sub=Sum('subtotal'), vat=Sum('vat_amount'), total=Sum('total_amount'), paid=Sum('paid_amount'))
    page_obj, num_pages = _paginate(qs, page)
    return render(request, 'reports/report_invoices.html', {
        'page_obj': page_obj, 'num_pages': num_pages, 'page': page,
        'totals': totals, 'date_from': date_from, 'date_to': date_to,
        'status': status, 'STATUS_MAP': STATUS_MAP,
    })


# ─── 2. VAT ────────────────────────────────────────────────────────────────
@login_required(login_url=LOGIN_URL)
def report_vat(request):
    from invoices.models import Invoice
    date_from, date_to = _date_filters(request)
    export = request.GET.get('export', '')
    page   = int(request.GET.get('page', 1))

    qs = Invoice.objects.exclude(status='draft').select_related('client').order_by('-issue_date')
    if date_from: qs = qs.filter(issue_date__gte=date_from)
    if date_to:   qs = qs.filter(issue_date__lte=date_to)

    if export == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'ضريبة القيمة المضافة'
        _style_header(ws, ['#','رقم الفاتورة','العميل','التاريخ','الوعاء الضريبي','نسبة الضريبة %','قيمة الضريبة','الإجمالي'])
        for i, inv in enumerate(qs, 1):
            ws.append([i, inv.invoice_number, inv.client.name if inv.client else '',
                str(inv.issue_date), float(inv.subtotal), float(inv.vat_rate), float(inv.vat_amount), float(inv.total_amount)])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 18
        return _excel_response(wb, 'تقرير_ضريبة_القيمة_المضافة')

    totals = qs.aggregate(sub=Sum('subtotal'), vat=Sum('vat_amount'), total=Sum('total_amount'))
    page_obj, num_pages = _paginate(qs, page)
    return render(request, 'reports/report_vat.html', {
        'page_obj': page_obj, 'num_pages': num_pages, 'page': page,
        'totals': totals, 'date_from': date_from, 'date_to': date_to,
    })


# ─── 3. REVENUE BY MONTH ────────────────────────────────────────────────────
@login_required(login_url=LOGIN_URL)
def report_revenue(request):
    from invoices.models import Invoice
    from django.db.models.functions import TruncMonth
    date_from, date_to = _date_filters(request)
    export = request.GET.get('export', '')

    qs = Invoice.objects.exclude(status__in=['draft','cancelled'])
    if date_from: qs = qs.filter(issue_date__gte=date_from)
    if date_to:   qs = qs.filter(issue_date__lte=date_to)

    monthly = (qs.annotate(month=TruncMonth('issue_date'))
                 .values('month').annotate(
                     total=Sum('total_amount'), vat=Sum('vat_amount'),
                     paid=Sum('paid_amount'),   cnt=Count('id'))
                 .order_by('month'))

    if export == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'الإيرادات الشهرية'
        _style_header(ws, ['الشهر','عدد الفواتير','إجمالي الإيرادات','الضريبة','المحصّل'])
        for row in monthly:
            ws.append([row['month'].strftime('%Y/%m') if row['month'] else '', row['cnt'],
                float(row['total'] or 0), float(row['vat'] or 0), float(row['paid'] or 0)])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 20
        return _excel_response(wb, 'تقرير_الإيرادات_الشهرية')

    total_all = qs.aggregate(t=Sum('total_amount'), v=Sum('vat_amount'), p=Sum('paid_amount'), c=Count('id'))
    return render(request, 'reports/report_revenue.html', {
        'monthly': list(monthly), 'total_all': total_all,
        'date_from': date_from, 'date_to': date_to,
    })


# ─── 4. CLIENTS ──────────────────────────────────────────────────────────────
@login_required(login_url=LOGIN_URL)
def report_clients(request):
    from clients.models import Client
    search = request.GET.get('q', '')
    export = request.GET.get('export', '')
    page   = int(request.GET.get('page', 1))

    clients = Client.objects.annotate(
        invoice_count=Count('invoices', filter=~Q(invoices__status='draft')),
        total_invoiced=Sum('invoices__total_amount', filter=~Q(invoices__status='draft')),
        total_paid_amt=Sum('invoices__paid_amount',  filter=~Q(invoices__status='draft')),
    ).order_by('-total_invoiced')
    if search: clients = clients.filter(Q(name__icontains=search) | Q(phone__icontains=search))

    if export == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'تقرير العملاء'
        _style_header(ws, ['#','الاسم','الهاتف','البريد','عدد الفواتير','إجمالي الفواتير','المحصّل','المتبقي'])
        for i, c in enumerate(clients, 1):
            ti = float(c.total_invoiced or 0); tp = float(c.total_paid_amt or 0)
            ws.append([i, c.name, c.phone or '', c.email or '', c.invoice_count or 0, ti, tp, ti-tp])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 18
        return _excel_response(wb, 'تقرير_العملاء')

    page_obj, num_pages = _paginate(clients, page)
    return render(request, 'reports/report_clients.html', {
        'page_obj': page_obj, 'num_pages': num_pages, 'page': page, 'search': search,
    })


# ─── 5. EMPLOYEES ────────────────────────────────────────────────────────────
@login_required(login_url=LOGIN_URL)
def report_employees(request):
    from employees.models import Employee
    search = request.GET.get('q', '')
    export = request.GET.get('export', '')
    page   = int(request.GET.get('page', 1))

    emps = Employee.objects.order_by('first_name', 'last_name')
    if search: emps = emps.filter(Q(first_name__icontains=search) | Q(last_name__icontains=search) | Q(job_title__icontains=search))

    if export == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'تقرير الموظفين'
        _style_header(ws, ['#','الاسم','المسمى الوظيفي','الهاتف','البريد'])
        for i, e in enumerate(emps, 1):
            ws.append([i, str(e), e.job_title or '', e.phone_number or '', e.email or ''])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 22
        return _excel_response(wb, 'تقرير_الموظفين')

    page_obj, num_pages = _paginate(emps, page)
    return render(request, 'reports/report_employees.html', {
        'page_obj': page_obj, 'num_pages': num_pages, 'page': page, 'search': search,
    })


# ─── 6. INVENTORY ────────────────────────────────────────────────────────────
@login_required(login_url=LOGIN_URL)
def report_inventory(request):
    from inventory.models import Item
    search   = request.GET.get('q', '')
    low_only = request.GET.get('low', '')
    export   = request.GET.get('export', '')
    page     = int(request.GET.get('page', 1))

    items = Item.objects.select_related('category').order_by('name')
    if search:   items = items.filter(Q(name__icontains=search) | Q(sku__icontains=search))
    if low_only: items = items.filter(qty_on_hand__lte=F('reorder_level'))

    if export == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'تقرير المخزون'
        _style_header(ws, ['#','اسم الصنف','الكود','الفئة','الكمية','مستوى إعادة الطلب','الوحدة','سعر الوحدة','القيمة'])
        for i, item in enumerate(items, 1):
            qty  = float(item.qty_on_hand or 0)
            price = float(item.unit_cost or 0)
            ws.append([i, item.name, item.sku or '', item.category.name if item.category else '',
                qty, float(item.reorder_level or 0), item.unit or '', price, qty*price])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 16
        return _excel_response(wb, 'تقرير_المخزون')

    page_obj, num_pages = _paginate(items, page)
    return render(request, 'reports/report_inventory.html', {
        'page_obj': page_obj, 'num_pages': num_pages, 'page': page,
        'search': search, 'low_only': low_only,
    })


# ─── 7. MARKETING ────────────────────────────────────────────────────────────
@login_required(login_url=LOGIN_URL)
def report_marketing(request):
    from marketing.models import MonthlyPlan, ExternalMarketer
    status = request.GET.get('status', '')
    export = request.GET.get('export', '')
    page   = int(request.GET.get('page', 1))
    STATUS_MAP = {'draft':'مسودة','approved':'معتمدة','completed':'مكتملة'}

    plans = MonthlyPlan.objects.prefetch_related('budgets').order_by('-month')
    if status: plans = plans.filter(status=status)

    if export == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'الخطط التسويقية'
        _style_header(ws, ['#','العنوان','الشهر','الحالة','إجمالي الميزانية'])
        for i, p in enumerate(plans, 1):
            ws.append([i, p.title, p.month.strftime('%Y/%m'), STATUS_MAP.get(p.status,''), float(p.total_budget)])
        ws2 = wb.create_sheet('المسوقون')
        _style_header(ws2, ['#','الاسم','الهاتف','البريد','العمولة %','جدول التسوية','الحالة'])
        for i, m in enumerate(ExternalMarketer.objects.all(), 1):
            ws2.append([i, m.name, m.phone or '', m.email or '', float(m.commission_rate), m.get_settlement_schedule_display(), 'نشط' if m.is_active else 'موقوف'])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 20
        for col in ws2.columns: ws2.column_dimensions[col[0].column_letter].width = 20
        return _excel_response(wb, 'تقرير_التسويق')

    marketers = ExternalMarketer.objects.all()
    page_obj, num_pages = _paginate(plans, page)
    return render(request, 'reports/report_marketing.html', {
        'page_obj': page_obj, 'num_pages': num_pages, 'page': page,
        'marketers': marketers, 'status': status, 'STATUS_MAP': STATUS_MAP,
    })


# ─── 8. RECEIPTS ─────────────────────────────────────────────────────────────
@login_required(login_url=LOGIN_URL)
def report_receipts(request):
    from invoices.models import PaymentReceipt
    date_from, date_to = _date_filters(request)
    method = request.GET.get('method', '')
    export = request.GET.get('export', '')
    page   = int(request.GET.get('page', 1))

    qs = PaymentReceipt.objects.select_related('invoice__client').order_by('-payment_date')
    if date_from: qs = qs.filter(payment_date__gte=date_from)
    if date_to:   qs = qs.filter(payment_date__lte=date_to)
    if method:    qs = qs.filter(payment_method=method)

    if export == 'excel':
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'سندات القبض'
        _style_header(ws, ['#','رقم السند','رقم الفاتورة','العميل','التاريخ','طريقة الدفع','المبلغ'])
        for i, r in enumerate(qs, 1):
            ws.append([i, r.receipt_number, r.invoice.invoice_number if r.invoice else '',
                r.invoice.client.name if r.invoice and r.invoice.client else '',
                str(r.payment_date), r.payment_method, float(r.amount)])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 18
        return _excel_response(wb, 'تقرير_سندات_القبض')

    totals = qs.aggregate(total=Sum('amount'), cnt=Count('id'))
    page_obj, num_pages = _paginate(qs, page)
    return render(request, 'reports/report_receipts.html', {
        'page_obj': page_obj, 'num_pages': num_pages, 'page': page,
        'totals': totals, 'date_from': date_from, 'date_to': date_to, 'method': method,
    })


# ─── 9. FINANCIAL STATEMENTS ─────────────────────────────────────────────────
@login_required(login_url=LOGIN_URL)
def report_financial(request):
    from decimal import Decimal
    from accounting.models import Account, AccountType, JournalLine
    from invoices.models import Invoice, PaymentReceipt

    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    export    = request.GET.get('export', '')

    def _account_balance(category, date_from='', date_to=''):
        """Sum of all posted journal lines for accounts of a given category."""
        qs = JournalLine.objects.filter(
            journal__is_posted=True,
            account__account_type__category=category,
        )
        if date_from: qs = qs.filter(journal__date__gte=date_from)
        if date_to:   qs = qs.filter(journal__date__lte=date_to)
        r = qs.aggregate(d=Sum('debit'), c=Sum('credit'))
        debit  = r['d'] or Decimal('0')
        credit = r['c'] or Decimal('0')
        # revenue & liability normal balance = credit; asset & expense = debit
        if category in ('revenue', 'liability', 'equity'):
            return credit - debit
        return debit - credit

    def _accounts_breakdown(category, date_from='', date_to=''):
        """Per-account breakdown for a category."""
        accounts = Account.objects.filter(
            account_type__category=category,
            is_active=True,
        ).prefetch_related('journal_lines__journal').order_by('code')
        rows = []
        for acc in accounts:
            qs = acc.journal_lines.filter(journal__is_posted=True)
            if date_from: qs = qs.filter(journal__date__gte=date_from)
            if date_to:   qs = qs.filter(journal__date__lte=date_to)
            r = qs.aggregate(d=Sum('debit'), c=Sum('credit'))
            debit  = r['d'] or Decimal('0')
            credit = r['c'] or Decimal('0')
            if category in ('revenue', 'liability', 'equity'):
                balance = credit - debit
            else:
                balance = debit - credit
            if balance != 0:
                rows.append({'code': acc.code, 'name': acc.name, 'balance': balance})
        return rows

    # ── Income Statement ──
    revenue_rows  = _accounts_breakdown('revenue',  date_from, date_to)
    expense_rows  = _accounts_breakdown('expense',  date_from, date_to)
    total_revenue = sum(r['balance'] for r in revenue_rows)
    total_expense = sum(r['balance'] for r in expense_rows)
    net_income    = total_revenue - total_expense

    # ── Balance Sheet ──
    asset_rows     = _accounts_breakdown('asset',     date_from, date_to)
    liability_rows = _accounts_breakdown('liability', date_from, date_to)
    equity_rows    = _accounts_breakdown('equity',    date_from, date_to)
    total_assets      = sum(r['balance'] for r in asset_rows)
    total_liabilities = sum(r['balance'] for r in liability_rows)
    total_equity      = sum(r['balance'] for r in equity_rows)

    # ── Invoice revenue summary (from invoices module) ──
    inv_qs = Invoice.objects.exclude(status__in=['draft', 'cancelled'])
    if date_from: inv_qs = inv_qs.filter(issue_date__gte=date_from)
    if date_to:   inv_qs = inv_qs.filter(issue_date__lte=date_to)
    inv_totals = inv_qs.aggregate(
        revenue=Sum('subtotal'),
        vat=Sum('vat_amount'),
        gross=Sum('total_amount'),
        collected=Sum('paid_amount'),
    )
    inv_outstanding = (inv_totals['gross'] or 0) - (inv_totals['collected'] or 0)

    # ── Cash collected (receipts) ──
    rec_qs = PaymentReceipt.objects.filter()
    if date_from: rec_qs = rec_qs.filter(payment_date__gte=date_from)
    if date_to:   rec_qs = rec_qs.filter(payment_date__lte=date_to)
    total_collected = rec_qs.aggregate(t=Sum('amount'))['t'] or Decimal('0')

    if export == 'excel':
        wb = openpyxl.Workbook()

        # Sheet 1 — Income Statement
        ws1 = wb.active; ws1.title = 'قائمة الدخل'
        _style_header(ws1, ['الحساب', 'اسم الحساب', 'المبلغ'])
        ws1.append(['── الإيرادات ──', '', ''])
        for r in revenue_rows:
            ws1.append([r['code'], r['name'], float(r['balance'])])
        ws1.append(['', 'إجمالي الإيرادات', float(total_revenue)])
        ws1.append(['── المصروفات ──', '', ''])
        for r in expense_rows:
            ws1.append([r['code'], r['name'], float(r['balance'])])
        ws1.append(['', 'إجمالي المصروفات', float(total_expense)])
        ws1.append(['', 'صافي الدخل', float(net_income)])
        for col in ws1.columns: ws1.column_dimensions[col[0].column_letter].width = 28

        # Sheet 2 — Balance Sheet
        ws2 = wb.create_sheet('الميزانية العمومية')
        _style_header(ws2, ['الحساب', 'اسم الحساب', 'المبلغ'])
        ws2.append(['── الأصول ──', '', ''])
        for r in asset_rows:
            ws2.append([r['code'], r['name'], float(r['balance'])])
        ws2.append(['', 'إجمالي الأصول', float(total_assets)])
        ws2.append(['── الخصوم ──', '', ''])
        for r in liability_rows:
            ws2.append([r['code'], r['name'], float(r['balance'])])
        ws2.append(['', 'إجمالي الخصوم', float(total_liabilities)])
        ws2.append(['── حقوق الملكية ──', '', ''])
        for r in equity_rows:
            ws2.append([r['code'], r['name'], float(r['balance'])])
        ws2.append(['', 'إجمالي حقوق الملكية', float(total_equity)])
        for col in ws2.columns: ws2.column_dimensions[col[0].column_letter].width = 28

        # Sheet 3 — Invoice Revenue Summary
        ws3 = wb.create_sheet('ملخص إيرادات الفواتير')
        _style_header(ws3, ['البند', 'المبلغ'])
        ws3.append(['إجمالي الإيرادات (قبل الضريبة)', float(inv_totals['revenue'] or 0)])
        ws3.append(['ضريبة القيمة المضافة', float(inv_totals['vat'] or 0)])
        ws3.append(['الإجمالي شامل الضريبة', float(inv_totals['gross'] or 0)])
        ws3.append(['المحصّل', float(inv_totals['collected'] or 0)])
        ws3.append(['المتبقي', float(inv_outstanding)])
        for col in ws3.columns: ws3.column_dimensions[col[0].column_letter].width = 36

        return _excel_response(wb, 'القوائم_المالية')

    return render(request, 'reports/report_financial.html', {
        'revenue_rows':      revenue_rows,
        'expense_rows':      expense_rows,
        'total_revenue':     total_revenue,
        'total_expense':     total_expense,
        'net_income':        net_income,
        'asset_rows':        asset_rows,
        'liability_rows':    liability_rows,
        'equity_rows':       equity_rows,
        'total_assets':      total_assets,
        'total_liabilities': total_liabilities,
        'total_equity':      total_equity,
        'inv_totals':        inv_totals,
        'inv_outstanding':   inv_outstanding,
        'total_collected':   total_collected,
        'date_from':         date_from,
        'date_to':           date_to,
    })
